import csv
import io
import logging
import re
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
from datetime import datetime
from .validation import validate_drill_score, get_unit_for_drill, DRILL_SCORE_RANGES
from ..services.schema_registry import SchemaRegistry

logger = logging.getLogger(__name__)

class ImportResult:
    def __init__(self, valid_rows: List[Dict[str, Any]], errors: List[Dict[str, Any]], detected_sport: str = "unknown", confidence: str = "low", sheets: List[Dict[str, Any]] = None):
        self.valid_rows = valid_rows
        self.errors = errors
        self.detected_sport = detected_sport
        self.confidence = confidence
        self.sheets = sheets or []

class DataImporter:
    """
    Utility class to parse and normalize input data (CSV, Excel, Text)
    into a standardized list of player objects with drill results.
    Supports multi-sport schema detection.
    """
    
    REQUIRED_HEADERS = ['first_name', 'last_name']
    OPTIONAL_HEADERS = ['age_group', 'jersey_number', 'email', 'phone', 'position']
    
    # Map common variations to canonical field names
    FIELD_MAPPING = {
        'first': 'first_name',
        'fname': 'first_name',
        'firstname': 'first_name',
        'last': 'last_name',
        'lname': 'last_name',
        'lastname': 'last_name',
        'jersey': 'jersey_number',
        'number': 'jersey_number',
        'no': 'jersey_number',
        '#': 'jersey_number',
        'age': 'age_group',
        'group': 'age_group',
        'division': 'age_group',
        'pos': 'position',
        'bib': 'external_id',
        'bib_number': 'external_id',
        'bib_no': 'external_id',
        'bib_#': 'external_id',
        'bibno': 'external_id',
        'external_id': 'external_id',
        'athlete_id': 'external_id'
    }

    @staticmethod
    def _normalize_header(header: str, schema_drills: List[str] = None, drill_label_map: Dict[str, str] = None) -> str:
        """
        Normalize header string to match canonical field names or schema drill keys.
        
        Args:
            header: Raw header from CSV/Excel
            schema_drills: List of valid drill keys
            drill_label_map: Dict mapping normalized labels to drill keys (e.g., {"bench_press": "x7hG4kL9mN2pQ8vW"})
        
        Returns:
            Canonical field name or drill key
        """
        if not header:
            return ""
        
        # Remove units in parentheses (e.g., "Lane Agility (sec)" -> "Lane Agility")
        header_no_units = re.sub(r'\s*\([^)]*\)\s*', ' ', str(header))
        clean = header_no_units.strip().lower().replace(' ', '_').replace('-', '_')
        
        # Check exact matches first
        if clean in DataImporter.FIELD_MAPPING:
            return DataImporter.FIELD_MAPPING[clean]
        
        # CRITICAL FIX: Check if normalized header matches a drill label
        if drill_label_map and clean in drill_label_map:
            return drill_label_map[clean]
            
        # Check if it matches any known schema drill keys (if provided)
        if schema_drills and clean in schema_drills:
            return clean
            
        # Check if it matches any known legacy football drill keys
        if clean in DRILL_SCORE_RANGES:
            return clean
            
        # Fuzzy Matching Logic
        
        # Prioritize Specific Compounds that might overlap with generic terms
        if 'free' in clean and 'throw' in clean: return 'free_throws'
        if 'exit' in clean and 'vel' in clean: return 'exit_velocity'
        
        # Basketball - check these before football to avoid conflicts
        if 'lane' in clean and 'agil' in clean: return 'lane_agility'
        
        # Robust 3-point matching (must come before fuzzy matching)
        if clean in ("3_point", "3pt", "three_point", "3_point_made", "3pt_made", "three_point_made", "3_point_shooting", "3_pt", "3_pt_made"):
            return "three_point"
        if "3_point" in clean or "three_point" in clean or "3pt" in clean or "3_pt" in clean:
            return "three_point"
            
        if ('3pt' in clean or '3_pt' in clean or 'three_point' in clean or '3_point' in clean) and 'shoot' in clean: return 'three_point'
        if '3_point' in clean or 'three_point' in clean: return 'three_point' # Catch-all for "3 Point"
        if 'spot' in clean and 'shoot' in clean: return 'three_point'  # "Spot Shooting" -> three_point
        
        # Football
        if '40' in clean and 'dash' in clean: return '40m_dash'
        if 'jump' in clean or 'vert' in clean: return 'vertical_jump'
        if 'catch' in clean: return 'catching'
        if 'throw' in clean and 'vel' not in clean and 'free' not in clean: return 'throwing' # Avoid overlap with throwing_velocity and free_throws
        if 'agil' in clean and 'lane' not in clean: return 'agility' # Avoid overlap with lane_agility
        
        # Baseball
        if 'pop' in clean: return 'pop_time'
        if 'fielding' in clean: return 'fielding_accuracy'
        
        # Basketball (already checked above for lane_agility and three_point)
        if 'dribble' in clean or 'handl' in clean or 'ball_handl' in clean: return 'dribbling'
        if 'defensive' in clean and 'slide' in clean: return 'defensive_slide'
        
        # Soccer
        if 'ball' in clean and 'control' in clean: return 'ball_control'
        if 'pass' in clean: return 'passing_accuracy'
        if 'shoot' in clean and 'power' in clean: return 'shooting_power'
        
        # Track
        if '100' in clean: return 'sprint_100'
        if '400' in clean: return 'sprint_400'
        if 'long' in clean and 'jump' in clean: return 'long_jump'
        if 'shot' in clean: return 'shot_put'
        if 'mile' in clean: return 'mile_time'
        
        # Volleyball
        if 'approach' in clean: return 'approach_jump'
        if 'serve' in clean: return 'serving_accuracy'
        if 'block' in clean: return 'blocking_reach'
            
        return clean

    @staticmethod
    def _clean_value(value: Any) -> Optional[float]:
        """
        Smart Error Correction: Clean and normalize numeric values.
        Handles:
        - Units (s, sec, in, ", etc.)
        - European decimals (4,52 -> 4.52)
        - Typos (4..5 -> 4.5)
        - Whitespace
        """
        if value is None:
            return None
            
        s_val = str(value).strip().lower()
        if not s_val:
            return None
            
        # Remove common units
        s_val = re.sub(r'[a-z"%]+$', '', s_val).strip() # Remove trailing units like 's', 'in', '"', '%'
        
        # Replace comma with dot (European decimal)
        s_val = s_val.replace(',', '.')
        
        # Fix double dots (typo)
        s_val = s_val.replace('..', '.')
        
        try:
            return float(s_val)
        except ValueError:
            return None

    @staticmethod
    def _detect_sport(headers: List[str]) -> Tuple[str, str]:
        """
        Auto-detect sport type based on headers.
        Returns (sport_id, confidence)
        """
        normalized = [DataImporter._normalize_header(h) for h in headers]
        schemas = SchemaRegistry.get_all_schemas()
        
        best_sport = "football" # Default
        max_score = 0
        
        for schema in schemas:
            score = 0
            drill_keys = [d.key for d in schema.drills]
            for h in normalized:
                if h in drill_keys:
                    score += 1
            
            # Normalize score by number of drills to avoid bias toward larger schemas
            # But prefer higher absolute matches too.
            if score > max_score:
                max_score = score
                best_sport = schema.id
        
        if max_score >= 3:
            return best_sport, "high"
        elif max_score >= 1:
            return best_sport, "medium"
            
        return "football", "low"

    @staticmethod
    def parse_csv(content: bytes, event_id: str = None, disabled_drills: List[str] = None) -> ImportResult:
        """Parse CSV content"""
        try:
            # Decode bytes to string
            text = content.decode('utf-8-sig') # Handle BOM if present
            f = io.StringIO(text)
            reader = csv.DictReader(f)
            
            # Normalize headers
            if not reader.fieldnames:
                return ImportResult([], [{"row": 0, "message": "Empty CSV file"}])
                
            # Detect Sport
            sport, confidence = DataImporter._detect_sport(reader.fieldnames)
            
            # CRITICAL FIX: If event_id provided, use event schema (includes custom drills)
            # Otherwise fall back to base sport schema
            if event_id:
                from ..utils.event_schema import get_event_schema
                schema = get_event_schema(event_id)
            else:
                schema = SchemaRegistry.get_schema(sport)
            
            schema_drills = [d.key for d in schema.drills] if schema else []
            
            # CRITICAL FIX: Build drill label to key mapping for custom drills
            drill_label_map = {}
            if schema:
                for drill in schema.drills:
                    normalized_label = drill.label.strip().lower().replace(' ', '_').replace('-', '_')
                    normalized_key = drill.key.strip().lower().replace(' ', '_').replace('-', '_')
                    # Map label to key if they're different
                    if normalized_label != normalized_key:
                        drill_label_map[normalized_label] = drill.key
            
            normalized_field_map = {
                field: DataImporter._normalize_header(field, schema_drills, drill_label_map) 
                for field in reader.fieldnames
            }
            
            result = DataImporter._process_rows(reader, normalized_field_map, sport, event_id, disabled_drills)
            result.detected_sport = sport
            result.confidence = confidence
            return result
            
        except Exception as e:
            logger.error(f"CSV Parse Error: {e}")
            return ImportResult([], [{"row": 0, "message": f"Failed to parse CSV: {str(e)}"}])

    @staticmethod
    def parse_excel(content: bytes, sheet_name: Optional[str] = None, event_id: str = None, disabled_drills: List[str] = None) -> ImportResult:
        """
        Parse Excel (XLSX) content.
        If multiple sheets exist and no sheet_name provided, returns list of sheets.
        """
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            
            # Handle multi-sheet detection
            if not sheet_name and len(wb.sheetnames) > 1:
                sheets_info = []
                for name in wb.sheetnames:
                    ws = wb[name]
                    # Get first 3 rows for preview
                    preview = []
                    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
                        preview.append([str(cell or "") for cell in row])
                    sheets_info.append({
                        "name": name,
                        "preview": preview
                    })
                return ImportResult([], [], sheets=sheets_info)
            
            # Select worksheet
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    return ImportResult([], [{"row": 0, "message": f"Sheet '{sheet_name}' not found"}])
            else:
                ws = wb.active
            
            # Read rows
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return ImportResult([], [{"row": 0, "message": "Empty Excel sheet"}])
                
            # Extract headers from first row
            headers = [str(cell or "").strip() for cell in rows[0]]
            
            # Detect Sport
            sport, confidence = DataImporter._detect_sport(headers)
            
            # CRITICAL FIX: If event_id provided, use event schema (includes custom drills)
            # Otherwise fall back to base sport schema
            if event_id:
                from ..utils.event_schema import get_event_schema
                schema = get_event_schema(event_id)
            else:
                schema = SchemaRegistry.get_schema(sport)
            
            schema_drills = [d.key for d in schema.drills] if schema else []
            
            # CRITICAL FIX: Build drill label to key mapping for custom drills
            drill_label_map = {}
            if schema:
                for drill in schema.drills:
                    normalized_label = drill.label.strip().lower().replace(' ', '_').replace('-', '_')
                    normalized_key = drill.key.strip().lower().replace(' ', '_').replace('-', '_')
                    if normalized_label != normalized_key:
                        drill_label_map[normalized_label] = drill.key
            
            # Map headers
            normalized_field_map = {
                headers[i]: DataImporter._normalize_header(headers[i], schema_drills, drill_label_map)
                for i in range(len(headers))
            }
            
            # Convert to dict list for processing
            data_rows = []
            for row_idx, row in enumerate(rows[1:], start=2):
                row_data = {}
                has_data = False
                for col_idx, cell_value in enumerate(row):
                    if col_idx < len(headers):
                        key = headers[col_idx]
                        if cell_value is not None and str(cell_value).strip() != "":
                            row_data[key] = cell_value
                            has_data = True
                if has_data:
                    data_rows.append(row_data)
            
            result = DataImporter._process_rows(data_rows, normalized_field_map, sport, event_id, disabled_drills)
            result.detected_sport = sport
            result.confidence = confidence
            return result
            
        except Exception as e:
            logger.error(f"Excel Parse Error: {e}")
            return ImportResult([], [{"row": 0, "message": f"Failed to parse Excel file: {str(e)}"}])

    @staticmethod
    def parse_image(content: bytes, event_id: str = None, disabled_drills: List[str] = None) -> ImportResult:
        """
        Parse image content using OCR.
        """
        try:
            from .ocr import OCRProcessor
            
            lines, confidence = OCRProcessor.extract_rows_from_image(content)
            
            if not lines:
                 return ImportResult([], [{"row": 0, "message": "No text detected in image"}])
                 
            # Convert lines to CSV-like string for parsing
            csv_text = OCRProcessor.lines_to_csv_string(lines)
            
            # Reuse parse_text logic
            result = DataImporter.parse_text(csv_text, event_id=event_id, disabled_drills=disabled_drills)
            
            # Override confidence if needed, but for now rely on structure detection
            return result
            
        except ImportError:
            logger.error("OCR dependencies not installed")
            return ImportResult([], [{"row": 0, "message": "OCR system not initialized"}])
        except Exception as e:
            logger.error(f"Image Parse Error: {e}")
            return ImportResult([], [{"row": 0, "message": f"Failed to parse image: {str(e)}"}])

    @staticmethod
    def parse_text(text: str, event_id: str = None, disabled_drills: List[str] = None) -> ImportResult:
        """
        Parse pasted text. Assumes either CSV-like structure or specific format.
        For now, implements a robust delimiter sniffer (tab, comma, pipe).
        """
        try:
            # Trim and split lines
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            if not lines:
                return ImportResult([], [{"row": 0, "message": "No text provided"}])
                
            # Sniff delimiter from header row
            header_line = lines[0]
            delimiters = ['\t', ',', '|', ';']
            best_delimiter = ','
            max_cols = 0
            
            for d in delimiters:
                cols = len(header_line.split(d))
                if cols > max_cols:
                    max_cols = cols
                    best_delimiter = d
            
            # Parse using CSV reader with detected delimiter
            f = io.StringIO(text)
            reader = csv.DictReader(f, delimiter=best_delimiter)
            
            if not reader.fieldnames:
                return ImportResult([], [{"row": 0, "message": "Could not parse headers"}])

            # Detect Sport
            sport, confidence = DataImporter._detect_sport(reader.fieldnames)
            
            # CRITICAL FIX: If event_id provided, use event schema (includes custom drills)
            # Otherwise fall back to base sport schema
            if event_id:
                from ..utils.event_schema import get_event_schema
                schema = get_event_schema(event_id)
            else:
                schema = SchemaRegistry.get_schema(sport)
            
            schema_drills = [d.key for d in schema.drills] if schema else []

            # CRITICAL FIX: Build drill label to key mapping for custom drills
            drill_label_map = {}
            if schema:
                for drill in schema.drills:
                    normalized_label = drill.label.strip().lower().replace(' ', '_').replace('-', '_')
                    normalized_key = drill.key.strip().lower().replace(' ', '_').replace('-', '_')
                    if normalized_label != normalized_key:
                        drill_label_map[normalized_label] = drill.key

            normalized_field_map = {
                field: DataImporter._normalize_header(field, schema_drills, drill_label_map) 
                for field in reader.fieldnames
            }
            
            result = DataImporter._process_rows(reader, normalized_field_map, sport, event_id, disabled_drills)
            result.detected_sport = sport
            result.confidence = confidence
            return result
            
        except Exception as e:
            logger.error(f"Text Parse Error: {e}")
            return ImportResult([], [{"row": 0, "message": f"Failed to parse text: {str(e)}"}])

    @staticmethod
    def _process_rows(rows: Any, field_map: Dict[str, str], sport_id: str, event_id: Optional[str] = None, disabled_drills: List[str] = None) -> ImportResult:
        """Common processing logic for all input types"""
        valid_rows = []
        errors = []
        
        # Load Schema for Validation
        # CRITICAL FIX: If event_id is provided, use get_event_schema to include CUSTOM DRILLS
        # Otherwise fall back to static template registry
        if event_id:
            from ..utils.event_schema import get_event_schema
            schema = get_event_schema(event_id)
        else:
            schema = SchemaRegistry.get_schema(sport_id)

        if not schema:
            # Fallback to football if detection failed completely
            schema = SchemaRegistry.get_schema("football")
            
        # Identify which columns map to drill scores from Schema
        # NOTE: If using get_event_schema(), disabled drills are already filtered out
        # The disabled_drills parameter is only needed for base schema fallback
        drill_keys = set(d.key for d in schema.drills)
        
        # FILTER DISABLED DRILLS (only if using base schema, not event schema)
        # Custom drills should NEVER be filtered - if created, they should be used
        if disabled_drills and not event_id:
            # Only filter when using base schema (no event_id means base schema)
            drill_keys = drill_keys - set(disabled_drills)
            
        drill_defs = {d.key: d for d in schema.drills}
        
        for idx, row in enumerate(rows, start=1):
            processed_row = {}
            row_errors = []
            found_canonical_keys = set()
            
            # Map fields
            for original_key, value in row.items():
                mapped_key = field_map.get(original_key)
                if not mapped_key:
                    continue
                    
                clean_val = str(value).strip() if value is not None else ""
                
                if mapped_key in drill_keys and clean_val:
                    found_canonical_keys.add(mapped_key)
                    # SMART ERROR CORRECTION: Try to fix common formatting issues
                    cleaned_num = DataImporter._clean_value(clean_val)
                    drill_def = drill_defs.get(mapped_key)
                    
                    if cleaned_num is not None:
                        # Validate Range from Schema
                        if drill_def:
                            min_v = drill_def.min_value if drill_def.min_value is not None else -1000
                            max_v = drill_def.max_value if drill_def.max_value is not None else 10000
                            if not (min_v <= cleaned_num <= max_v):
                                row_errors.append(f"Value {cleaned_num} for '{original_key}' out of range ({min_v}-{max_v})")
                                processed_row[original_key] = cleaned_num # Keep value to show error
                            else:
                                processed_row[original_key] = cleaned_num
                        else:
                            processed_row[original_key] = cleaned_num
                    else:
                        row_errors.append(f"Invalid number format for '{original_key}': '{clean_val}'")
                        processed_row[original_key] = clean_val # Keep raw value
                
                elif mapped_key == 'jersey_number' and clean_val:
                    found_canonical_keys.add(mapped_key)
                    try:
                        num = int(float(clean_val)) # Handle "10.0" from Excel
                        processed_row[original_key] = num
                    except ValueError:
                        row_errors.append(f"Invalid player number: {clean_val}")
                        processed_row[original_key] = clean_val
                
                else:
                    # Regular string fields
                    if mapped_key:
                        found_canonical_keys.add(mapped_key)
                    
                    if clean_val:
                        processed_row[original_key] = clean_val

            # Check required fields
            if 'first_name' not in found_canonical_keys:
                row_errors.append("Missing First Name")
            if 'last_name' not in found_canonical_keys:
                row_errors.append("Missing Last Name")
                
            # Construct result item
            item = {
                "row_id": idx,
                "data": processed_row,
                "errors": row_errors,
                "original": str(row) # Debug helper
            }
            
            if row_errors:
                errors.append({
                    "row": idx,
                    "data": processed_row,
                    "message": "; ".join(row_errors)
                })
            else:
                valid_rows.append(item)
                
        return ImportResult(valid_rows, errors)
